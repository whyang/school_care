# -*- coding: utf-8 -*-
"""
Created on Mon Jan. 7, 2022
@author: Wen-Hsin Yang

說明：
使用 pandas, openpyxl, pywin32整理安親班學生收費資料，製作學生收費工作表及收費單(Excel worksheet, PDF file) 
1. 讀入[安親斑_收費資料.xlsx]: 名冊、收費項目、收費期間 等內容
2. 整理成[收費單.xlsx]：每位學生使用1個工作表
3. 由[收費單.xlsx] 製作 [收費單_合併.xlsx]、[收費單_合併.pdf]：可列印用收費單(Excel、PDF)
"""
import os
import pandas as pd
import numpy as np
import win32com.client
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
def student_list():
    ###
    # Function: 根據「收費資料」內容，組出「收費單_名冊」
    ##
    
    ##
    # 1. 讀入收費資料(學生名單、收費項目、收費月份、註記事項)
    global df_student, df_item, df_month, df_note # set as global variables
    with pd.ExcelFile('安親班_收費資料.xlsx') as xlsxFile:
        df_student = pd.read_excel(xlsxFile, sheet_name='學生名單', header=0)
        df_item = pd.read_excel(xlsxFile, sheet_name='收費項目', header=0) 
        df_month = pd.read_excel(xlsxFile, sheet_name='收費月份', header=0)
        df_note = pd.read_excel(xlsxFile, sheet_name='註記事項', header=0)
    ##
    # 2. 轉收費單表格(每位學生一張)
    df = df_month.T # 轉置'月份'(由一欄轉為一列)
    df.columns = df.iloc[0] # 將月份設定成資料框表頭(欄位名稱)
    df.drop(df.index[0], inplace=True) # 移除原先第一列'月份'各欄位 
    df = pd.concat([df_item, df], axis=1) # 加入'收費項目'欄，合併月份資料框
    #df.fillna(int(0), inplace=True) # 資料框中的cell的空值，用'整數0'取代

    ##
    # 3. 加入學生名單於收費單表格(每位學生一張)；每位學生一個Excel工作表(所有的收費項目)
    name_list = df_student['姓名']
    df_note_temp = pd.DataFrame()
    df_note_temp['項目/月份'] = df_note['註記'] # 將註記內容放在['項目/月份'](註記位置對齊收費單第一欄)
    with pd.ExcelWriter('收費單_名冊.xlsx') as writer:
        for idx, student in enumerate(name_list, start=0):
            # 將註記內容放在['項目/月份'](註記位置對齊收費單第一欄)
            # 收費單(資料框)第一欄名稱為'項目/月份'
            df_temp = pd.concat([df, df_note_temp], axis=0) # 合併收費表單、註記內容
            
            #將每一個學生的收費單轉成numpy array (context)
            content = df_temp.to_numpy() #轉成numpy array        
            # 放回第一列，放入原來資料框欄位名稱
            # 因為資料框轉成numpy arrany後，少了原先欄位名稱(column lable)這一列
            content = np.insert(content, 0, [df_temp.columns], axis=0) # 加入'欄位名稱'到第一列  
            
            # 將每個學生的收費表，加入表頭(學生年級、班級、姓名)
            content = np.insert(content, 0, [np.NAN], axis=0) # 插入新的第一列(空白列)
            # 取回每位學生對應屬於那一個年級與班級
            tableTitle = df_student.iloc[idx]['年級'] + '年' +  df_student.iloc[idx]['班級'] + '班  ' + student
            content[0, 0] = tableTitle
            
            # 產出 收費單_名冊 excel worksheet，將 NAN 用 ''取代
            df_temp = pd.DataFrame(content) #轉成資料框(dataframe) (學生收費單名冊)
            # 產生工作表名稱(年級+班級+_+姓名)
            sheetName = df_student.iloc[idx]['年級'] + df_student.iloc[idx]['班級'] + '_' + student
            df_temp.to_excel(writer, sheet_name=sheetName, na_rep='', header=False, index=False) # cell空值 NAN 用 ''取代
            #df_temp.to_excel(writer, sheet_name=student, na_rep='', header=False, index=False) # cell空值 NAN 用 ''取代
#
# end of student_list()
##

def single_payment_list(row_base=0, sourceFile='收費單_名冊.xlsx', targetFile='收費單.xlsx'):
    ###
    # Function: 根據「收費單_名冊」，製作每位學生之收費單表格(Excel檔案，每位學生一張工作表)
    # Parameter:
    #   row_base: 開始列編號(表格開始，因為由1開始，所以base號碼為0)
    #   sourceFile: 讀入的收費單名冊資料(每位學生的原始收費資料)
    #   targetFile: 產生每位學生的收費單(每位學生1張工作表)
    ##
    name_list = df_student['姓名'] # 取出所有收費學生姓名清單
    with pd.ExcelFile(sourceFile) as xlsxFile:
        wb = load_workbook(xlsxFile)
        #for student in name_list:
        for idx, student in enumerate(name_list, start=0):
            # 將學生姓名對應到收費單名冊上工作表名稱
            sheetName = df_student.iloc[idx]['年級'] + df_student.iloc[idx]['班級'] + '_' + student
            ##
            # 讀入每一個學生的收費單(one excel's worksheet)       
            #ws = wb[student]
            ws = wb[sheetName]

            ##
            # 合併第一列(年級, 班級, 姓名)儲存格 (根據工作表的最大寬度_欄位數目)
            i = row_base + 1
            for idx in range(1, ws.max_column+1, 1):
                ws.cell(row=i, column=idx).border = style_head['border']
                ws.cell(row=i, column=idx).fill = style_head['fill']
                ws.cell(row=i, column=idx).font = style_head['font']
                ws.cell(row=i, column=idx).alignment = style_head['alignment']
            ws.row_dimensions[i].height = 20 #30
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=ws.max_column)               
        
            ##    
            # 畫出表格(收費表格)，設定格式(style)
            row, col = df_item.shape #df_month.shape # 取出 收費表格 的大小(長、寬)(col, row)
            row += 1 # 因為df_item 資料框不含表頭(欄位名稱)，但是收費表格中"欄位名稱"需要顯示為一列
            for j in range(2, row+2, 1): # row 需加上第一列，且range計數結束條件值方式，所以row加上2
            #for j in range(2, row+2, 1): # row 需加上第一列，且range計數結束條件值方式，所以row加上2
                r = row_base + j
                for k in range(1, ws.max_column+1, 1): # 設定每個儲存格(cell)的格式
                    ws.cell(row=r, column=k).border = style_content['border']
                    ws.cell(row=r, column=k).font = style_content['font']
                    if k == 1:
                        ws.cell(row=r, column=k).alignment = style_content['alignment']
                    else:
                        ws.cell(row=r, column=k).alignment = style_head['alignment']
                # 設定表格儲存格的大小
                ws.row_dimensions[r].height = 18 #22 # 列的固定高度
                ws.column_dimensions['A'].width = 15 # 第一欄('項目/日期')固定寬度
                # 調整各月份儲存格的寬度
                if ws.max_column >= 2 and ws.max_column <= 7:
                    col_dic = col_obj[ws.max_column] # 取出對應的字典(dictionary object)物件
                    for idx in range(2, ws.max_column+1, 1):
                        col_idx_alphabet = str(col_idx[idx]) # 取出對應到工作表欄位名稱(i.e., 'A','B',...)
                        ws.column_dimensions[col_idx_alphabet].width = col_dic[idx] # 取出儲存格(月份)對應的寬度大小
        
            ##
            # 註記事項 
            for idx in range(row+2, ws.max_row+1, 1): # row 需加上第1列，且range計數起始值設成1開始，所以row加上2
                i = row_base + idx
                ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=ws.max_column)
                ws.cell(row=i, column=1).font = style_note['font'] #style_content['font']
                ws.cell(row=i, column=1).alignment = style_note['alignment'] #style_content['alignment']
                ws.row_dimensions[i].height = 15
        ##
        # 整理所有學生收費單(worksheet)後，儲存到收費單檔案(Excel)
        wb.save(targetFile)
        wb.close()     
#
# end of single_payment_list()
##

def convert_to_pdf(sourceFile='收費單_合併.xlsx', targetFile='收費單.pdf'):
    ###
    # Function: 將Excel工作表，轉換成PDF格式，以方便列印
    # Parameter:
    #   filename: 檔案名稱
    ##       
    
    ##
    # pip install --upgrade pywin32==225
    # 配合 Python 3.8 版本，使用pywin32 version 225
    o = win32com.client.Dispatch("Excel.Application") # 引用pywin32 package
    o.Visible = True
    
    ##
    # 準備資料: 
    # sourceFile(Excel檔案名稱)
    # targetFile(PDF檔案名稱)
    dirpath = os.getcwd()
    wb_p_path = os.path.join(dirpath, sourceFile) # source filename with path info.
    path_to_pdf = os.path.join(dirpath, targetFile) # target filename with path info.
    # 將收費單工作表轉成序列號碼
    ws_index_list = []
    for i, sheetName in enumerate(merged_sheetName, start=1):
        ws_index_list.append(i)
    
    ##
    # 讀取Excel工作表
    wb_p = o.Workbooks.Open(wb_p_path)   
    wb_p.Sheets(ws_index_list).Select()
    
    ##
    # Excel工作表轉成PDF格式檔案
    wb_p.ActiveSheet.ExportAsFixedFormat(0, path_to_pdf)
    
    wb_p.Close(True) 
    o.Quit()
#
# end of convert_to_pdf()
##

def merged_payment_list(noMerge=3, sourceFile='收費單.xlsx', targetFile='收費單_合併.xlsx'):
    ###
    # Function: 根據「收費單」，製作合併學生收費單表格(Excel檔案，幾位學生合成一張工作表)，做為列印用
    # Parameter:
    #   noMerge: 合併表格(收費單)數目
    #   sourceFile: 讀入的收費單資料(每位學生的收費表)
    #   targetFile: 產生幾位學生合併的收費單(幾位學生1張工作表)
    ##       
    wb_source = load_workbook(sourceFile) # 所有學生收費單
    wb_target_book = Workbook() # 合併合併學生收費單表格 (Excel檔案，noMerge個學生收費單合成一張工作表)
    
    name_list = df_student['姓名'] # 取出所有收費學生姓名清單
    # 將學生姓名對應到收費單名冊上工作表名稱
    sheetName_list = []
    for idx, student in enumerate(name_list, start=0):
        sheetName = df_student.iloc[idx]['年級'] + df_student.iloc[idx]['班級'] + '_' + student
        sheetName_list.append(sheetName)
    
    length = len(sheetName_list) # 學生數(收費單數目)，依照收費單工作表(取代姓名) 
    itemRow, itemCol = df_item.shape # number of the table of the item (row, column)
    
    global merged_sheetName
    merged_sheetName = [] # record the whole sheet names after merging

    ##
    # 根據 noMerge 值，分批(noMerge+1個新的工作表)產生合併的收費單(工作表)
    base = 0
    for i in range(length // noMerge):
        base = i * noMerge # base index of worksheet
        rowIdx = 0 # initial row index of the new worksheet 
        
        # 合併 noMerge 個工作表
        for idx in range(noMerge):
            sheetName = sheetName_list[base + idx] #name_list[base+idx]
            ws = wb_source[sheetName] # 每位學生收費單(工作表)
            if idx == 0:
                sheetName_new = '收費單_列印_'+str(i+1) 
                merged_sheetName.append(sheetName_new)
                wb_target_book.create_sheet(sheetName_new)
                ws_target =  wb_target_book[sheetName_new]

            # copy each row in worksheet to the merged target worksheet     
            for index_r, row in enumerate(ws.rows, start=1):
                r = rowIdx + index_r
                for index_c, col in enumerate(row, start=1):
                    cell = col.value
                    ws_target.cell(row=r, column=index_c).value = cell
                ###
                # editing table
                if index_r == 1:
                    # merge cells in according to each sheet of the tablet
                    # 年級, 班級, 姓名
                    ws_target.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ws.max_column) 
                    cellIdx = 'A' + str(r)
                    #ws_target[cellIdx].border = style_head['border']
                    ws_target[cellIdx].fill = style_head['fill']
                    ws_target[cellIdx].font = style_head['font']
                    ws_target[cellIdx].alignment = style_head['alignment']    
                    ws_target.row_dimensions[r].height = 20 #30
                elif (index_r >= 2) and (index_r <= itemRow+2): # '收費項目'數字+1(表格之表頭佔1列)，再+1(加上學生年級、班、姓名)
                    # 畫收費表格
                    for k in range(1, ws.max_column+1, 1): # 設定每個儲存格(cell)的格式
                        ws_target.cell(row=r, column=k).border = style_content['border']
                        ws_target.cell(row=r, column=k).font = style_content['font']
                        if k == 1:
                            ws_target.cell(row=r, column=k).alignment = style_content['alignment']
                        else:
                            ws_target.cell(row=r, column=k).alignment = style_head['alignment']
                    # 設定表格儲存格的大小
                    ws_target.row_dimensions[r].height = 18 #22 # 列的固定高度
                    ws_target.column_dimensions['A'].width = 15 # 第一欄('項目/日期')固定寬度
                    # 調整各月份儲存格的寬度
                    if ws.max_column >= 2 and ws.max_column <= 7:
                        col_dic = col_obj[ws.max_column] # 取出對應的字典(dictionary object)物件
                        for idx in range(2, ws.max_column+1, 1):
                            col_idx_alphabet = str(col_idx[idx]) # 取出對應到工作表欄位名稱(i.e., 'A','B',...)
                            ws_target.column_dimensions[col_idx_alphabet].width = col_dic[idx] # 取出儲存格(月份)對應的寬度大小       
                else:
                    # 註記事項
                    ws_target.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ws.max_column)
                    ws_target.cell(row=r, column=1).font = style_note['font'] #style_content['font']
                    ws_target.cell(row=r, column=1).alignment = style_note['alignment'] #style_content['alignment']
                    ws_target.row_dimensions[r].height = 15                            
            # update row index
            rowIdx = r + 2 # 加入2列空白於兩個學生收費單之間
        
    base = (i+1)*noMerge # 調整到下一批合併收費單的起始點
    rowIdx = 0 # row index of worksheet

    ##    
    # 根據 noMerge 值，檢查分批(noMerge個新的合併工作表後)，如還有剩下的收費單合併成最後1個工作表
    for idx in range(length % noMerge):
        sheetName = sheetName_list[base + idx] #name_list[base + idx]
        ws = wb_source[sheetName]
        if idx == 0:
            sheetName_new = '收費單_列印_'+str(i+2) 
            merged_sheetName.append(sheetName_new)
            wb_target_book.create_sheet(sheetName_new)
            ws_target =  wb_target_book[sheetName_new]
        # copy each row in worksheet to the target merged workshee     
        for index_r, row in enumerate(ws.rows, start=1):
            r = rowIdx + index_r
            for index_c, col in enumerate(row, start=1):
                x1 = col.value
                ws_target.cell(row=r, column=index_c).value = x1

            ###
            # editing table
            if index_r == 1:
                # merge cells according to the table of each sheet
                # 年級, 班級, 姓名
                ws_target.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ws.max_column) 
                cellIdx = 'A' + str(r)
                #ws_target[cellIdx].border = style_head['border']
                ws_target[cellIdx].fill = style_head['fill']
                ws_target[cellIdx].font = style_head['font']
                ws_target[cellIdx].alignment = style_head['alignment']    
                ws_target.row_dimensions[r].height = 20 #30
            elif (index_r >= 2) and (index_r <= itemRow+2): # '收費項目'數字+1(表格之表頭佔1列)，再+1(加上學生年級、班、姓名)
                # 畫收費表格
                for k in range(1, ws.max_column+1, 1): # 設定每個儲存格(cell)的格式
                    ws_target.cell(row=r, column=k).border = style_content['border']
                    ws_target.cell(row=r, column=k).font = style_content['font']
                    if k == 1:
                        ws_target.cell(row=r, column=k).alignment = style_content['alignment']
                    else:
                        ws_target.cell(row=r, column=k).alignment = style_head['alignment']
                # 設定表格儲存格的大小
                ws_target.row_dimensions[r].height = 18 #22 # 列的固定高度
                ws_target.column_dimensions['A'].width = 15 # 第一欄('項目/日期')固定寬度
                # 調整各月份儲存格的寬度
                if ws.max_column >= 2 and ws.max_column <= 7:
                    col_dic = col_obj[ws.max_column] # 取出對應的字典(dictionary object)物件
                    for idx in range(2, ws.max_column+1, 1):
                        col_idx_alphabet = str(col_idx[idx]) # 取出對應到工作表欄位名稱(i.e., 'A','B',...)
                        ws_target.column_dimensions[col_idx_alphabet].width = col_dic[idx] # 取出儲存格(月份)對應的寬度大小 
            else:
                # 註記事項
                ws_target.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ws.max_column)
                ws_target.cell(row=r, column=1).font = style_note['font'] #style_content['font']
                ws_target.cell(row=r, column=1).alignment = style_note['alignment'] #style_content['alignment']
                ws_target.row_dimensions[r].height = 15 
        
        # update row index
        rowIdx = r + 2 # 加入2列空白於兩個學生收費單之間 
    
    ##
    # finally, 產出合併收費單(Excel worksheet以及PDF)
    wb_target_book.remove(wb_target_book['Sheet']) # 移除建立workbook時，有一個預設的'Sheet'工作表
    wb_target_book.save(targetFile) # 產出合併收費單(Excel worksheet)
    wb_target_book.close()
    wb_source.close()
    #convert_to_pdf(sourceFile='收費單_合併.xlsx', targetFile='收費單_合併.pdf') # 產出合併收費單(PDF)
    convert_to_pdf(sourceFile='收費單_合併.xlsx', targetFile='收費單.pdf') # 產出合併收費單(PDF)
#
# end of merged_payment_list()
##

if __name__ == '__main__':  
    ##
    # 宣告使用的格式:
    # 定義表頭的樣式
    style_head = {
        "border": Border(left=Side(style='medium', color='FF000000'), 
                         right=Side(style='medium', color='FF000000'),
                         top=Side(style='medium', color='FF000000'), 
                         bottom=Side(style='medium', color='FF000000')),
        "fill": PatternFill("solid", fgColor="9AFF9A"),
        "font": Font(color="000000", bold=True, name="標楷體", size=14),
        "alignment": Alignment(horizontal="center", vertical="center")
        }
    # 定義表格內容樣式
    style_content = {
        "border": Border(left=Side(style='thin', color='FF000000'), 
                         right=Side(style='thin', color='FF000000'),
                         top=Side(style='thin', color='FF000000'), 
                         bottom=Side(style='thin', color='FF000000')),
        "alignment": Alignment(horizontal='left', vertical='center'),
        "font": Font(name="標楷體", size=14)}
    # 定義表格儲存格的大小，2個欄位~7個欄位
    columns_2 = {1:15, 2:60}
    columns_3 = {1:15, 2:30, 3:30}
    columns_4 = {1:15, 2:20, 3:20, 4:20}
    columns_5 = {1:15, 2:15, 3:15, 4:15, 5:15}
    columns_6 = {1:15, 2:12, 3:12, 4:12, 5:12, 6:12}
    columns_7 = {1:15, 2:10, 3:10, 4:10, 5:10, 6:10, 7:10}
    col_idx = {1:'A', 2:'B', 3:'C', 4:'D', 5:'E', 6:'F', 7:'G'} # 欄位名稱(對應工作表 A~G)
    col_obj = {2:columns_2, 3:columns_3, 4:columns_4, 5:columns_5, 6:columns_6, 7:columns_7} #對應取出字典物件(dictionary object)名稱
    
    # 定義備註樣式
    style_note = {
        "border": Border(left=Side(style='thin', color='FF000000'), 
                         right=Side(style='thin', color='FF000000'),
                         top=Side(style='thin', color='FF000000'), 
                         bottom=Side(style='thin', color='FF000000')),
        "alignment": Alignment(horizontal='left', vertical='center'),
        "font": Font(name="標楷體", size=10)}
    
    ##
    # 根據收費資料，製作學生收費表(Excel工作表)、列印用收費單(Excel、PDF)
    # 1. 收費資料轉成收費表框架
    student_list() 
    # 2. 製作每位學生的收費表(Excel工作表)
    single_payment_list(sourceFile='收費單_名冊.xlsx', targetFile='收費單.xlsx') 
    # 3. 製作列印用收費表(Excel工作表、PDF)
    merged_payment_list(noMerge=3, sourceFile='收費單.xlsx', targetFile='收費單_合併.xlsx') 

    ##
    # 刪除不必要的檔案
    file_1 = '收費單_名冊.xlsx'
    file_2 = '收費單_合併.xlsx'
    dirpath = os.getcwd()
    file_1_path = os.path.join(dirpath, file_1) # filename with path info.
    file_2_path = os.path.join(dirpath, file_2) # filename with path info.
    try:
        os.remove(file_1_path)
        os.remove(file_2_path)
    except OSError as e:
        print(e)
    else:
        pass
    
    ##
    # ending
    print('好了! 請看 收費單.xlsx 收費單.pdf')
#
# end of main()
##