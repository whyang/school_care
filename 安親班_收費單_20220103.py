# -*- coding: utf-8 -*-
"""
Created on Fri Dec 24 16:35:13 2021

@author: whyang
"""

###
# 程式 1
##
import pandas as pd
import numpy as np

with pd.ExcelFile('安親班_收費資料.xlsx') as xlsxFile:
    df_student = pd.read_excel(xlsxFile, sheet_name='學生名單', header=0)
    df_item = pd.read_excel(xlsxFile, sheet_name='收費項目', header=0) 
    df_month = pd.read_excel(xlsxFile, sheet_name='收費月份', header=0)
    df_note = pd.read_excel(xlsxFile, sheet_name='註記事項', header=0)

df = df_month.T
df.columns = df.iloc[0]
df.drop(df.index[0], inplace=True)

df = pd.concat([df_item, df], axis=1)
#df.fillna(int(0), inplace=True)

# 收費單 名冊
# 每位學生一個Excel工作表(所有的收費項目)
name_list = df_student['姓名']
i = 0
df_note_temp = pd.DataFrame()
with pd.ExcelWriter('安親班_收費單_名冊.xlsx') as writer:
    for student in name_list:
        #將註記的內容，放在column['項目/月份'] (放在收費單的第一欄位，以對齊收費單和註記的位置)
        #收費單(資料框dataframe:df) 的第一個欄位,其名稱為'項目/月份'
        df_note_temp['項目/月份'] = df_note['註記']
        df_temp = pd.concat([df, df_note_temp], axis=0)
        #產出 收費單 excel worksheet
        df_temp.to_excel(writer, sheet_name=student, index=False)
        i += 1

# 收費單 列印
# 每位學生一個Excel工作表 (列印收費單)
i = 0
with pd.ExcelWriter('安親班_收費單_列印.xlsx') as writer:
    for student in name_list:
        #讀入每一個學生的收費單(excel worksheet)
        df_tempt = pd.read_excel('安親班_收費單_名冊.xlsx', sheet_name=student, header=None)

        #將每一個學生的收費單轉成numpy array (context)
        content = df_temp.to_numpy() #轉成numpy array
        
        #放回一列：欄位名稱
        #因為 用 with..as..，包納read_excel後，header=None的功能 disable
        #所以轉成numpy arrany後少了原先欄位名稱(column lable)這一列
        content1 = np.insert(content, 0, [df_temp.columns], axis=0)
        
        #將每一個學生的收費表，加入表頭(學生年級、班級、姓名)
        content2 = np.insert(content1, 0, [np.NAN], axis=0) # numpy array加入第一列(表頭)
        content2[0, 0] = df_student.iloc[i]['年級'] + '年' +  df_student.iloc[i]['班級'] + '班  ' + student
        
        df2 = pd.DataFrame(content2) #轉成資料框(dataframe) (列印之學生收費單)
        
        #產出 列印收費單 excel worksheet，將 NAN 用 ''取代 
        df2.to_excel(writer, sheet_name=student, na_rep='', header=False, index=False)
        i += 1

###
# 程式 2
##        
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

with pd.ExcelFile('安親班_收費資料.xlsx') as xlsxFile:
    df_student = pd.read_excel(xlsxFile, sheet_name='學生名單', header=0)
    df_month = pd.read_excel(xlsxFile, sheet_name='收費月份', header=0)
 
# 定義表頭的樣式
style_head = {
    "border": Border(left=Side(style='medium', color='FF000000'), 
                     right=Side(style='medium', color='FF000000'),
                     top=Side(style='medium', color='FF000000'), 
                     bottom=Side(style='medium', color='FF000000')),
    "fill": PatternFill("solid", fgColor="9AFF9A"),
    "font": Font(color="000000", bold=True, name="標楷體", size=14),
    "alignment": Alignment(horizontal="center", vertical="center")
}

# 定義表內容樣式
style_content = {
    "border": Border(left=Side(style='thin', color='FF000000'), 
                     right=Side(style='thin', color='FF000000'),
                     top=Side(style='thin', color='FF000000'), 
                     bottom=Side(style='thin', color='FF000000')),
    "alignment": Alignment(horizontal='left', vertical='center'),
    "font": Font(name="標楷體")}

name_list = df_student['姓名']
with pd.ExcelFile('安親班_收費單_列印.xlsx') as xlsxFile:
    wb = load_workbook(xlsxFile)
    for student in name_list:
        #讀入每一個學生的收費單(excel worksheet)       
        ws = wb[student]
        
        cols = []
        for col in ws.iter_cols():
            cols.append(col)
        rows = []
        for row in ws.iter_rows():
            rows.append(row)
        
        # merge cells according to the table of each sheet
        # 年級, 班級, 姓名
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))       
        #ws['A1'].border = style_head['border']
        ws['A1'].fill = style_head['fill']
        ws['A1'].font = style_head['font']
        ws['A1'].alignment = style_head['alignment']    
        ws.row_dimensions[1].height = 30
        
        row, col = df_month.shape
        
        # 畫表格
        for j in range(2, len(rows)-row+2):
            for k in range(len(cols)):
                ws.cell(row=j, column=k+1).border = style_head['border']
                ws.cell(row=j, column=k+1).font = style_content['font']
                ws.cell(row=j, column=k+1).alignment = style_content['alignment']
            ws.column_dimensions['A'].width = 31
            ws.row_dimensions[j].height = 20
            
        # 註記事項               
        for i in range(row-1):
            ws.merge_cells(start_row=len(rows)-i, start_column=1, end_row=len(rows)-i, end_column=len(cols))
            #ws.cell(row=len(rows)-i, column=1).border = style_content['border']
            ws.cell(row=len(rows)-i, column=1).font = style_content['font']
            ws.cell(row=len(rows)-i, column=1).alignment = style_content['alignment']
            ws.row_dimensions[len(rows)-i].height = 15
        
wb.save('收費單.xlsx') 
wb.close()

###
# convet to pdf format
##

import win32com.client
import os

o = win32com.client.Dispatch("Excel.Application")
o.Visible = True #False

dirpath = os.getcwd()
filename = '收費單.xlsx'
wb_p_path = os.path.join(dirpath, filename)

wb_p = o.Workbooks.Open(wb_p_path)
ws_index_list = []
for i, sheetName in enumerate(name_list, start=1):
    ws_index_list.append(i)

filename = '收費單.pdf'
path_to_pdf = os.path.join(dirpath, filename)

wb_p.Sheets(ws_index_list).Select()
wb_p.ActiveSheet.ExportAsFixedFormat(0, path_to_pdf)
wb_p.Close(True) 

###
# 刪除不必要的檔案
##

import os

file_1 = r'安親班_收費單_名冊.xlsx'
file_2 = r'安親班_收費單_列印.xlsx'

try:
    os.remove(file_1)
    os.remove(file_2)
except OSError as e:
    print(e)
else:
    print("Files are deleted successfully")
    
print('work down')