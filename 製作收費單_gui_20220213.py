# -*- coding: utf-8 -*-
"""
Created on Mon Jan. 7, 2022
@author: Wen-Hsin Yang

說明：
使用 pandas, openpyxl, pywin32整理安親班學生收費資料，製作學生收費工作表及收費單(Excel worksheet, PDF file) 
0. 使用 tkinter 做為圖形操作介面(GUI)
1. 讀入[安親斑_收費資料.xlsx]: 名冊、收費項目、收費期間 等內容
2. 整理成[收費單.xlsx]：每位學生使用1個工作表
3. 根據[收費單.xlsx] 製作成 [收費單_合併.xlsx]，最後產出[收費單.pdf]：可列印用收費單(PDF)
"""
import os
import sys
import pandas as pd
import numpy as np
import win32com.client

import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import tkinter as tk
from tkinter import ttk
from tkinter import filedialog as fd
import tkinter.font as tkFont
from tkinter.messagebox import showinfo

from PIL import Image, ImageTk
from datetime import datetime

#import traceback
    
def student_list(fileName, outputPath):
    def xls_getSheet(_wb, _sheetName):
        # 學生名單
        _ws = _wb.sheet_by_name(_sheetName)
        headers = [str(cell.value) for cell in _ws.row(0)]
        arr = []
        for rowind in range(_ws.nrows)[1:]:
            arr.append([ cell.value for cell in _ws.row(rowind)])
        _content = np.rec.fromrecords(arr, names=headers)
        return pd.DataFrame(_content)
           
    progress() # 更新執行進度(processbar進度條)
    
    try:    
        ###
        # Function: 根據「收費資料」內容，組出「收費單_名冊」
        ##
           
        ##
        # 1. 讀入收費資料(學生名單、收費項目、收費月份、註記事項)
        global df_student, df_item, df_month, df_note # set as global variables
    
        # 判斷檔案 xlsx、Xls
        _fileFormat = fileName.split('.')[-1]    
        if _fileFormat == 'xls': 
            #import xlrd
            wb = xlrd.open_workbook(fileName)
            # 學生名單
            df_student = xls_getSheet(wb, '學生名單')
            df_student.fillna(str('  '), inplace=True) # 資料框中的cell的空值，用'空白字串'取代
            # 收費項目
            df_item = xls_getSheet(wb, '收費項目')
            # 收費月份
            df_month = xls_getSheet(wb, '收費月份')
            # 註記事項
            df_note = xls_getSheet(wb, '註記事項')
        else:
            with pd.ExcelFile(fileName) as xlsxFile:
                df_student = pd.read_excel(xlsxFile, sheet_name='學生名單', header=0)
                df_student.fillna(str('  '), inplace=True) # 資料框中的cell的空值，用'空白字串'取代
                df_item = pd.read_excel(xlsxFile, sheet_name='收費項目', header=0)
                df_month = pd.read_excel(xlsxFile, sheet_name='收費月份', header=0)
                df_note = pd.read_excel(xlsxFile, sheet_name='註記事項', header=0)
    
        df_item.rename(columns={df_item.columns[0]: '項目/月份'}, inplace=True)
        df_month.rename(columns={df_month.columns[0]: '項目/月份'}, inplace=True)
        progress() # 更新執行進度(processbar進度條)
        
        ##
        # 2. 轉收費單表格(每位學生一張)
        df = df_month.T # 轉置'月份'(由一欄轉為一列)
        df.columns = df.iloc[0] # 將月份設定成資料框表頭(欄位名稱)
        df.drop(df.index[0], inplace=True) # 移除原先第一列'月份'各欄位 
        df = pd.concat([df_item, df], axis=1) # 加入'收費項目'欄，合併月份資料框
        #df.fillna(int(0), inplace=True) # 資料框中的cell的空值，用'整數0'取代
        
        ##
        # 3. 加入學生名單於收費單表格(每位學生一張)；每位學生一個Excel工作表(所有的收費項目)
        _name = df_student.columns[0]
        _grade = df_student.columns[1]
        _class = df_student.columns[2]           
        name_list = df_student[_name] # 取出所有學生姓名清單
        df_note_temp = pd.DataFrame()
        df_note_temp['項目/月份'] = df_note[df_note.columns[0]] # 將註記內容放在['項目/月份'](註記位置對齊收費單第一欄)
        outputFile = os.path.join(outputPath, '收費單_名冊.xlsx')
        with pd.ExcelWriter(outputFile) as writer:
            for idx, student in enumerate(name_list, start=0):
                # 將註記內容放在['項目/月份'](註記位置對齊收費單第一欄)
                # 收費單(資料框)第一欄名稱為'項目/月份'
                df_temp = pd.concat([df, df_note_temp], axis=0) # 合併收費表單、註記內容
                #將每一個學生的收費單轉成numpy array (context)
                content = df_temp.to_numpy() #轉成numpy array
                # 放回第一列，放入原來資料框欄位名稱
                # 因為資料框轉成numpy arrany後，少了原先欄位名稱(column lable)這一列
                content = np.insert(content, 0, [df_temp.columns], axis=0) # 加入'欄位名稱'到第一列            
                # 將每個學生的收費表，加入表頭(學生年級、班級、姓名)
                content = np.insert(content, 0, [np.NAN], axis=0) # 插入新的第一列(空白列)
                # 取回每位學生對應屬於那一個年級與班級
                tableTitle = df_student.iloc[idx][_grade] + '年' + df_student.iloc[idx][_class] + '班  ' + student                
                content[0, 0] = tableTitle
         
                # 產出 收費單_名冊 excel worksheet，將 NAN 用 ''取代
                df_temp = pd.DataFrame(content) #轉成資料框(dataframe) (學生收費單名冊)
                # 產生工作表名稱(年級+班級+_+姓名)
                sheetName = df_student.iloc[idx][_grade] + df_student.iloc[idx][_class] + '_' + student
                df_temp.to_excel(writer, sheet_name=sheetName, na_rep='', header=False, index=False) # cell空值 NAN 用 ''取代
        
        progress() # 更新執行進度(processbar進度條)       
        return True
    
    except:
        e = sys.exc_info()[0]
        print(e)
        showinfo(
            title = '讀取收費資料，製作收費清單過程 (student_list)', 
            message = 
            ' 收費資料 檔案讀取錯誤 !! 請檢查四個工作表單內容："學生名單"、"收費項目"、"收費月份"、"註記事項" 。')
        stop()
        return False
#
# end of student_list()
##

def single_payment_list(row_base=0, sourceFile='收費單_名冊.xlsx', targetFile='收費單.xlsx', outputPath='.\\'):
    progress() # 更新執行進度(processbar進度條)
    
    ###
    # Function: 根據「收費單_名冊」，製作每位學生之收費單表格(Excel檔案，每位學生一張工作表)
    # Parameter:
    #   row_base: 開始列編號(表格開始，因為由1開始，所以base號碼為0)
    #   sourceFile: 讀入的收費單名冊資料(每位學生的原始收費資料)
    #   targetFile: 產生每位學生的收費單(每位學生1張工作表)
    ##
    _name = df_student.columns[0]    
    _grade = df_student.columns[1]
    _class = df_student.columns[2]
    name_list = df_student[_name] # 取出所有收費學生姓名清單
    sourceFile = os.path.join(outputPath, sourceFile)
    with pd.ExcelFile(sourceFile) as xlsxFile:
        wb = load_workbook(xlsxFile)
        for idx, student in enumerate(name_list, start=0):
            # 將學生姓名對應到收費單名冊上工作表名稱
            sheetName = df_student.iloc[idx][_grade] + df_student.iloc[idx][_class] + '_' + student
            ##
            # 讀入每一個學生的收費單(one excel's worksheet)       
            ws = wb[sheetName]

            ##
            # 合併第一列(年級, 班級, 姓名)儲存格 (根據工作表的最大寬度_欄位數目)
            i = row_base + 1
            for idx in range(1, ws.max_column+1, 1):
                ws.cell(row=i, column=idx).border = style_head['border']
                ws.cell(row=i, column=idx).fill = style_head['fill']
                ws.cell(row=i, column=idx).font = style_head['font']
                ws.cell(row=i, column=idx).alignment = style_head['alignment']
            ws.row_dimensions[i].height = 20 #30
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=ws.max_column)               
        
            ##    
            # 畫出表格(收費表格)，設定格式(style)
            row, col = df_item.shape #df_month.shape # 取出 收費表格 的大小(長、寬)(col, row)
            row += 1 # 因為df_item 資料框不含表頭(欄位名稱)，但是收費表格中"欄位名稱"需要顯示為一列
            for j in range(2, row+2, 1): # row 需加上第一列，且range計數結束條件值方式，所以row加上2
            #for j in range(2, row+2, 1): # row 需加上第一列，且range計數結束條件值方式，所以row加上2
                r = row_base + j
                for k in range(1, ws.max_column+1, 1): # 設定每個儲存格(cell)的格式
                    ws.cell(row=r, column=k).border = style_content['border']
                    ws.cell(row=r, column=k).font = style_content['font']
                    if k == 1:
                        ws.cell(row=r, column=k).alignment = style_content['alignment']
                    else:
                        ws.cell(row=r, column=k).alignment = style_head['alignment']
                # 設定表格儲存格的大小
                ws.row_dimensions[r].height = 18 #22 # 列的固定高度
                ws.column_dimensions['A'].width = 15 # 第一欄('項目/日期')固定寬度
                # 調整各月份儲存格的寬度
                if ws.max_column >= 2 and ws.max_column <= 7:
                    col_dic = col_obj[ws.max_column] # 取出對應的字典(dictionary object)物件
                    for idx in range(2, ws.max_column+1, 1):
                        col_idx_alphabet = str(col_idx[idx]) # 取出對應到工作表欄位名稱(i.e., 'A','B',...)
                        ws.column_dimensions[col_idx_alphabet].width = col_dic[idx] # 取出儲存格(月份)對應的寬度大小
        
            ##
            # 註記事項 
            for idx in range(row+2, ws.max_row+1, 1): # row 需加上第1列，且range計數起始值設成1開始，所以row加上2
                i = row_base + idx
                ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=ws.max_column)
                ws.cell(row=i, column=1).font = style_note['font'] #style_content['font']
                ws.cell(row=i, column=1).alignment = style_note['alignment'] #style_content['alignment']
                ws.row_dimensions[i].height = 15
        ##
        # 整理所有學生收費單(worksheet)後，儲存到收費單檔案(Excel)
        targetFile = os.path.join(outputPath, targetFile)
        wb.save(targetFile)
        wb.close()  
        
    progress() # 更新執行進度(processbar進度條)
#
# end of single_payment_list()
##

def convert_to_pdf(sourceFile='收費單_合併.xlsx', targetFile='收費單.pdf', dirpath=os.getcwd()):
    progress() # 更新執行進度(processbar進度條)
    ###
    # Function: 將Excel工作表，轉換成PDF格式，以方便列印
    # Parameter:
    #   filename: 檔案名稱
    ##       
    
    ##
    # pip install --upgrade pywin32==225
    # 配合 Python 3.8 版本，使用pywin32 version 225
    o = win32com.client.Dispatch("Excel.Application") # 引用pywin32 package
    o.Visible = False #True
    
    ##
    # 準備資料: 
    # sourceFile(Excel檔案名稱)
    # targetFile(PDF檔案名稱)
    wb_p_path = os.path.join(dirpath, sourceFile) # source filename with path info.
    path_to_pdf = os.path.join(dirpath, targetFile) # target filename with path info.
    # 將收費單工作表轉成序列號碼
    ws_index_list = []
    for i, sheetName in enumerate(merged_sheetName, start=1):
        ws_index_list.append(i)
    
    ##
    # 讀取Excel工作表
    wb_p = o.Workbooks.Open(wb_p_path)   
    wb_p.Sheets(ws_index_list).Select()
    
    progress() # 更新執行進度(processbar進度條)
    
    ##
    # Excel工作表轉成PDF格式檔案
    wb_p.ActiveSheet.ExportAsFixedFormat(0, path_to_pdf)
    
    wb_p.Close(True) 
    o.Quit()
    progress() # 更新執行進度(processbar進度條)
#
# end of convert_to_pdf()
##

def merged_payment_list(noMerge=3, sourceFile='收費單.xlsx', targetFile='收費單_合併.xlsx', outputPath='.\\'):
    progress() # 更新執行進度(processbar進度條)
    
    ###
    # Function: 根據「收費單」，製作合併學生收費單表格(Excel檔案，幾位學生合成一張工作表)，做為列印用
    # Parameter:
    #   noMerge: 合併表格(收費單)數目
    #   sourceFile: 讀入的收費單資料(每位學生的收費表)
    #   targetFile: 產生幾位學生合併的收費單(幾位學生1張工作表)
    ##
    sourceFile = os.path.join(outputPath, sourceFile)   
    wb_source = load_workbook(sourceFile) # 所有學生收費單
    wb_target_book = Workbook() # 合併合併學生收費單表格 (Excel檔案，noMerge個學生收費單合成一張工作表)
    _name = df_student.columns[0]    
    _grade = df_student.columns[1]
    _class = df_student.columns[2]    
    name_list = df_student[_name] # 取出所有收費學生姓名清單
    # 將學生姓名對應到收費單名冊上工作表名稱
    sheetName_list = []
    for idx, student in enumerate(name_list, start=0):
        sheetName = df_student.iloc[idx][_grade] + df_student.iloc[idx][_class] + '_' + student
        sheetName_list.append(sheetName)
    
    length = len(sheetName_list) # 學生數(收費單數目)，依照收費單工作表(取代姓名) 
    itemRow, itemCol = df_item.shape # number of the table of the item (row, column)
    
    global merged_sheetName
    merged_sheetName = [] # record the whole sheet names after merging
    
    progress() # 更新執行進度(processbar進度條)
    
    ##
    # 根據 noMerge 值，分批(noMerge+1個新的工作表)產生合併的收費單(工作表)
    base = 0
    for i in range(length // noMerge):
        base = i * noMerge # base index of worksheet
        rowIdx = 0 # initial row index of the new worksheet 
        
        # 合併 noMerge 個工作表
        for idx in range(noMerge):
            sheetName = sheetName_list[base + idx] #name_list[base+idx]
            ws = wb_source[sheetName] # 每位學生收費單(工作表)
            if idx == 0:
                sheetName_new = '收費單_列印_'+str(i+1) 
                merged_sheetName.append(sheetName_new)
                wb_target_book.create_sheet(sheetName_new)
                ws_target =  wb_target_book[sheetName_new]

            # copy each row in worksheet to the merged target worksheet     
            for index_r, row in enumerate(ws.rows, start=1):
                r = rowIdx + index_r
                for index_c, col in enumerate(row, start=1):
                    cell = col.value
                    ws_target.cell(row=r, column=index_c).value = cell
                ###
                # editing table
                if index_r == 1:
                    # merge cells in according to each sheet of the tablet
                    # 年級, 班級, 姓名
                    ws_target.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ws.max_column) 
                    cellIdx = 'A' + str(r)
                    #ws_target[cellIdx].border = style_head['border']
                    ws_target[cellIdx].fill = style_head['fill']
                    ws_target[cellIdx].font = style_head['font']
                    ws_target[cellIdx].alignment = style_head['alignment']    
                    ws_target.row_dimensions[r].height = 20 #30
                elif (index_r >= 2) and (index_r <= itemRow+2): # '收費項目'數字+1(表格之表頭佔1列)，再+1(加上學生年級、班、姓名)
                    # 畫收費表格
                    for k in range(1, ws.max_column+1, 1): # 設定每個儲存格(cell)的格式
                        ws_target.cell(row=r, column=k).border = style_content['border']
                        ws_target.cell(row=r, column=k).font = style_content['font']
                        if k == 1:
                            ws_target.cell(row=r, column=k).alignment = style_content['alignment']
                        else:
                            ws_target.cell(row=r, column=k).alignment = style_head['alignment']
                    # 設定表格儲存格的大小
                    ws_target.row_dimensions[r].height = 18 #22 # 列的固定高度
                    ws_target.column_dimensions['A'].width = 15 # 第一欄('項目/日期')固定寬度
                    # 調整各月份儲存格的寬度
                    if ws.max_column >= 2 and ws.max_column <= 7:
                        col_dic = col_obj[ws.max_column] # 取出對應的字典(dictionary object)物件
                        for idx in range(2, ws.max_column+1, 1):
                            col_idx_alphabet = str(col_idx[idx]) # 取出對應到工作表欄位名稱(i.e., 'A','B',...)
                            ws_target.column_dimensions[col_idx_alphabet].width = col_dic[idx] # 取出儲存格(月份)對應的寬度大小       
                else:
                    # 註記事項
                    ws_target.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ws.max_column)
                    ws_target.cell(row=r, column=1).font = style_note['font'] #style_content['font']
                    ws_target.cell(row=r, column=1).alignment = style_note['alignment'] #style_content['alignment']
                    ws_target.row_dimensions[r].height = 15                            
            # update row index
            rowIdx = r + 1 # 加入1列空白於兩個學生收費單之間
        
    base = (i+1)*noMerge # 調整到下一批合併收費單的起始點
    rowIdx = 0 # row index of worksheet
    
    progress() # 更新執行進度(processbar進度條)
    
    ##    
    # 根據 noMerge 值，檢查分批(noMerge個新的合併工作表後)，如還有剩下的收費單合併成最後1個工作表
    for idx in range(length % noMerge):
        sheetName = sheetName_list[base + idx] #name_list[base + idx]
        ws = wb_source[sheetName]
        if idx == 0:
            sheetName_new = '收費單_列印_'+str(i+2) 
            merged_sheetName.append(sheetName_new)
            wb_target_book.create_sheet(sheetName_new)
            ws_target =  wb_target_book[sheetName_new]
        # copy each row in worksheet to the target merged workshee     
        for index_r, row in enumerate(ws.rows, start=1):
            r = rowIdx + index_r
            for index_c, col in enumerate(row, start=1):
                x1 = col.value
                ws_target.cell(row=r, column=index_c).value = x1

            ###
            # editing table
            if index_r == 1:
                # merge cells according to the table of each sheet
                # 年級, 班級, 姓名
                ws_target.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ws.max_column) 
                cellIdx = 'A' + str(r)
                #ws_target[cellIdx].border = style_head['border']
                ws_target[cellIdx].fill = style_head['fill']
                ws_target[cellIdx].font = style_head['font']
                ws_target[cellIdx].alignment = style_head['alignment']    
                ws_target.row_dimensions[r].height = 20 #30
            elif (index_r >= 2) and (index_r <= itemRow+2): # '收費項目'數字+1(表格之表頭佔1列)，再+1(加上學生年級、班、姓名)
                # 畫收費表格
                for k in range(1, ws.max_column+1, 1): # 設定每個儲存格(cell)的格式
                    ws_target.cell(row=r, column=k).border = style_content['border']
                    ws_target.cell(row=r, column=k).font = style_content['font']
                    if k == 1:
                        ws_target.cell(row=r, column=k).alignment = style_content['alignment']
                    else:
                        ws_target.cell(row=r, column=k).alignment = style_head['alignment']
                # 設定表格儲存格的大小
                ws_target.row_dimensions[r].height = 18 #22 # 列的固定高度
                ws_target.column_dimensions['A'].width = 15 # 第一欄('項目/日期')固定寬度
                # 調整各月份儲存格的寬度
                if ws.max_column >= 2 and ws.max_column <= 7:
                    col_dic = col_obj[ws.max_column] # 取出對應的字典(dictionary object)物件
                    for idx in range(2, ws.max_column+1, 1):
                        col_idx_alphabet = str(col_idx[idx]) # 取出對應到工作表欄位名稱(i.e., 'A','B',...)
                        ws_target.column_dimensions[col_idx_alphabet].width = col_dic[idx] # 取出儲存格(月份)對應的寬度大小 
            else:
                # 註記事項
                ws_target.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ws.max_column)
                ws_target.cell(row=r, column=1).font = style_note['font'] #style_content['font']
                ws_target.cell(row=r, column=1).alignment = style_note['alignment'] #style_content['alignment']
                ws_target.row_dimensions[r].height = 15 
        
        # update row index
        rowIdx = r + 1 # 加入1列空白於兩個學生收費單之間 
    
    progress() # 更新執行進度(processbar進度條)
    
    ##
    # finally, 產出合併收費單(Excel worksheet以及PDF)
    wb_target_book.remove(wb_target_book['Sheet']) # 移除建立workbook時，有一個預設的'Sheet'工作表
    targetFile = os.path.join(outputPath, targetFile)
    wb_target_book.save(targetFile) # 產出合併收費單(Excel worksheet)
    wb_target_book.close()
    wb_source.close()
    convert_to_pdf(sourceFile='收費單_合併.xlsx', targetFile='收費單.pdf', dirpath=outputPath) # 產出合併收費單(PDF)
#
# end of merged_payment_list()
##

def main_func():  
    progress() # 更新執行進度(processbar進度條)
    
    # 判斷收費資料 (輸入檔案) 是否存在，如不存在則用預先設定的檔案名稱    
    fileName = os.path.join(os.getcwd(), input_payer_info.get())
    if not os.path.isfile(fileName):
        fileName = os.path.join(os.getcwd(), '收費資料.xlsx')
    
    # 判斷製作的收費單檔案存放路徑 是否存在，如不存在則用default value
    outputPath = os.path.join(os.getcwd(), input_pay_list_path.get())
    try:
        os.makedirs(outputPath)
    except FileExistsError as e:   
        pass
    except OSError as e:
        outputPath = os.path.join(os.getcwd(), input_pay_list_path.get().split('\\')[-1])
        os.makedirs(outputPath)
        showinfo(
            title = '收費單存放目錄(main_func)',
            message = '輸入值錯誤 (請手動刪除根據"輸入值"自動產生(不必要)的空目錄)!! ' +
                        ' 建立收費單檔案存放目錄為 ' + outputPath)
    
    # 更新畫面上輸入值
    input_payer_info.set(fileName)
    input_pay_list_path.set(outputPath)
    window.update()
    
    # 根據收費資料，製作學生收費表(Excel工作表)、列印用收費單(Excel、PDF)
    # 1. 收費資料轉成收費表框架
    
    if student_list(fileName, outputPath): # successful for reading into 收費資料
        try:    
            # 2. 製作每位學生的收費表(Excel工作表)
            single_payment_list(sourceFile='收費單_名冊.xlsx', targetFile='收費單.xlsx', outputPath=outputPath)
            # 3. 製作列印用收費表(Excel工作表、PDF)
            merged_payment_list(noMerge=3, sourceFile='收費單.xlsx', targetFile='收費單_合併.xlsx', outputPath=outputPath)     
            ##
            # 刪除不必要的檔案
            file_1 = '收費單_名冊.xlsx'
            file_2 = '收費單_合併.xlsx'
            dirpath = outputPath #os.getcwd()
            file_1_path = os.path.join(dirpath, file_1) # filename with path info.
            file_2_path = os.path.join(dirpath, file_2) # filename with path info.
            os.remove(file_1_path)
            os.remove(file_2_path)
            
            progress() # 更新執行進度(processbar進度條)
            ##
            # 結束製作收費單功能
            showinfo(
                title = '製作收費清單(main_func)',
                message='收費單已經好了! 檔案 請查閱  ' + outputPath + '  收費單.xlsx、收費單.pdf')
            stop()
        except:
            e = sys.exc_info()[0]
            print(e)
            showinfo(
                title = '製作收費清單(main_func)',
                message='根據收費資料製作收費單過程錯誤! 請檢查。' + e)
            stop()
            window.destroy
    else:
        stop()
        window.destroy
#
# end of main_func()
##

##
# function handler for ProgressBar (進度軸)
# 
def update_progress_label():
    if pb['value'] > 100:
        pb['value'] = 100
    return f"執行進度： {pb['value']}%"

def progress():
    if pb['value'] < 100:
        pb['value'] += 8
        value_label['text'] = update_progress_label()
        window.update()
    else:
        pass
                 
def stop():
    pb.stop()
    value_label['text'] = update_progress_label()
#
# end of function handler for ProgressBar 進度軸)
## 

##
# function handler of filedialog
#
def select_file():
    filetypes = (
        ('xlsx', '*.xlsx'),
        ('xls', '*.xls'),
        ('All files', '*.*')
    )
    filename_current = input_payer_info.get()
    filename = fd.askopenfilename(
        title='收費資料',
        initialdir='.\\',
        filetypes=filetypes)
    
    if len(filename.strip()) > 0:
        input_payer_info.set(filename)
        showinfo(title='選取的檔案', message=filename)
    else:
        input_payer_info.set(filename_current)
        showinfo(title='選取的檔案', message=filename_current)

def select_directory():
    path_current = input_pay_list_path.get()
    _path = fd.askdirectory(parent=window, initialdir='.\\')
    if len(_path.strip()) > 0:
        input_pay_list_path.set(_path)
        showinfo(title='選取的目錄', message=_path)
    else:
        input_pay_list_path.set(path_current)
        showinfo(title='選取的目錄', message=path_current)
    
#
# end of function handler of filedialog
##

###
# 日期、時間更新顯示處理
##
def DateTime():
    now = datetime.now()
    data_time = now.strftime('%Y/%m/%d %H:%M:%S')
    return now, data_time

class Update_Clock():
    def __init__(self, rootframe):
        self.root = rootframe
        self.label = tk.Label(self.root, text='yy/mm/dd 00:00:00', bg='#DDDDD8', 
                              font=tkFont.Font(family="Arial", size=10))
        self.label.grid(column=3, row=0, sticky='e')
        self.update_clock()
        self.root.mainloop()
    def update_clock(self):
        now, data_time = DateTime()
        self.label.configure(text=data_time)
        self.root.after(1000, self.update_clock)
#
# end of 日期、時間更新顯示處理
###
        
def define_layout(obj, cols=1, rows=1):   
    def method(trg, col, row):
        for c in range(cols):    
            trg.columnconfigure(c, weight=1)
        for r in range(rows):
            trg.rowconfigure(r, weight=1)
    # end of method()

    if type(obj)==list:        
        [ method(trg, cols, rows) for trg in obj ]
    else:
        trg = obj
        method(trg, cols, rows)
# end of define_layout()

def GUI():
    ##
    # UI's constructor (adopt Tkinter)
    ##
    global window
    window = tk.Tk() # construct the root window
    #window.geometry('640x480')
    window.title('製作收費單')
    
    # configing parameter
    align_mode = 'nswe' # used for the layout of grid which represents central placement
    _padx = 10 # the separated padding space
    _pady = 10
    div_size = 200 # unit size of the division  
    fontStyle = tkFont.Font(family="微軟正黑體", size=16, weight='bold')
    fontStyle_0 = tkFont.Font(family="微軟正黑體", size=24, weight='bold')
    fontStyle_1 = tkFont.Font(family="微軟正黑體", size=12, weight='bold')
    fontStyle_2 = tkFont.Font(family="微軟正黑體", size=10)
    
    # allocate four frames (i.e., operating areas/status bar) 
    div_1 = tk.Frame(window, width=div_size*6 , height=div_size) # 標頭區(top/1st frame)
    div_2 = tk.Frame(window, width=div_size*6 , height=div_size , bg='green') # 輸入區(middle/2nd frame)
    div_3 = tk.Frame(window, width=div_size*6 , height=div_size) # , bg='blue') # 狀態區(progress bar; the up bottom/3rd frame)
    div_4 = tk.Frame(window, width=div_size*6, height=div_size, bg='#DDDDD8') # 訊息區(information bar; the down bottom/4th frame)

    # configure grid layout for the three operating areas (divisions)
    div_1.grid(column=0, row=0, padx=_padx, pady=_pady, columnspan=4, sticky=align_mode)
    div_2.grid(column=0, row=1, padx=_padx, pady=_pady, rowspan=2, columnspan=4, sticky=align_mode)
    div_3.grid(column=0, row=3, padx=_padx, pady=_pady, columnspan=4, sticky=align_mode)
    div_4.grid(column=0, row=4, padx=_padx, pady=_pady, columnspan=4, sticky=align_mode)

    # adjust the grid layout 
    define_layout(div_1, cols=4)
    define_layout(div_2, cols=4, rows=2)
    define_layout(div_3, cols=4)
    define_layout(div_4, cols=4) 
    
    ###
    # usage metaphore of the UI
    ##
    
    ##
    # 標頭區(top/1st frame)   
    # 顯示logo
    # 準備logo's image
    im = Image.open('./上華課照班LOGO.jpg')
    image_size = 70
    imTK = ImageTk.PhotoImage(im.resize((image_size+15, image_size)))
    # 將logo貼上
    title_logo = tk.Label(div_1, image=imTK)
    title_logo['height'] = image_size
    title_logo['width'] = image_size
    title_logo.grid(column=0, row=0, columnspan=1, sticky=align_mode) 
    # 顯示作業名稱
    title_lbl = tk.Label(div_1, text='製作收費單', bg='orange', fg='blue', font=fontStyle_0)
    title_lbl.grid(column=1, row=0, columnspan=2, sticky=align_mode)
    # 建立結束關閉作業功能
    title_button = tk.Button(div_1, text='結束', command=window.destroy, bg='red', fg='white', font=fontStyle)
    title_button.grid(column=3, row=0, padx=_padx, pady=_pady, sticky=tk.E)
    
    ##
    # 輸入區(middle/2nd frame)
    # row 1
    input1_lbl =tk.Label(div_2, text='收費資料 檔案名稱 ：', bg='green', fg='white', font=fontStyle_1)
    input1_lbl.grid(column=0, row=0, sticky=tk.E) #align_mode)
    global input_payer_info
    input_payer_info = tk.StringVar()
    input_payer_info.set('收費資料.xlsx')
    input1_entry = tk.Entry(div_2, width=50, textvariable=input_payer_info)
    input1_entry.grid(column=1, row=0, padx=_padx, pady=_pady, columnspan=2, sticky=tk.W)
    input1_button = ttk.Button(div_2, text='選擇', command=select_file)
    input1_button.grid(column=3, row=0, padx=_padx, pady=_pady, sticky=align_mode)
    
    # row 2
    input2_lbl = tk.Label(div_2, text='收費單 存放目錄 ：', bg='green', fg='white', font=fontStyle_1)
    input2_lbl.grid(column=0 , row=1, sticky=tk.E) #align_mode)
    global input_pay_list_path
    input_pay_list_path = tk.StringVar()
    # 抓取現在日期時間，做為收費單製作後的輸出目錄
    now, data_time = DateTime()
    #_date = now.strftime('%Y-%m-%d_%H%M%S')
    _date = now.strftime('%Y-%m-%d')
    input_pay_list_path.set('收費單_' + _date)
    input2_entry = tk.Entry(div_2, width=50, textvariable=input_pay_list_path)
    input2_entry.grid(column=1, row=1, padx=_padx, pady=_pady, columnspan=2, sticky=tk.W)
    input2_button = ttk.Button(div_2, text='選擇', command=select_directory)
    input2_button.grid(column=3, row=1, padx=_padx, pady=_pady, sticky=align_mode)
    
    ##
    # 狀態區(progress bar; the up bottom/3rd frame) 
    # progressbar (進度軸)
    global pb
    pb = ttk.Progressbar(div_3, orient='horizontal', mode='determinate', length=div_size*2) 
    pb.grid(column=1, row=0, padx=_padx, pady=_pady, columnspan=2)
    pb['value'] = 0
    global value_label
    # 顯示進度訊息
    value_label = ttk.Label(div_3, text=update_progress_label()) 
    value_label.grid(column=0, row=0, sticky=tk.W)
    # 開使按鍵
    progressbar_button = tk.Button(div_3, text='開始', command=main_func, bg='blue', fg='white', font=fontStyle)
    progressbar_button.grid(column=3, row=0, padx=_padx, pady=_pady, sticky=align_mode) 

    # 訊息區(information bar; the down bottom/4th frame)
    informationbar_lbl1 = tk.Label(div_4, text='版本 Ver. 2.2', bg='#DDDDD8', font=fontStyle_2) #, bg='yellow', fg='white')
    informationbar_lbl1.grid(column=0, row=0, sticky='news')
    informationbar_lbl2 = tk.Label(div_4, text='上華 課後照顧班', bg='#DDDDD8', fg ='#228822', font=fontStyle) #, bg='yellow', fg='white')
    informationbar_lbl2.grid(column=1, row=0, columnspan=2, sticky='news')
    # 日期、時間 定義宣告，移入Update_Clock class's constructor
    #informationbar_lbl3 = tk.Label(div_4, text='yy/mm/dd 00:00:00', bg='#DDDDD8', font=tkFont.Font(family="Arial", size=10)')
    #informationbar_lbl3.grid(column=3, row=0, sticky='e')
    Update_Clock(div_4)
    
    window.mainloop()
#
# end of GUI()
##
    
if __name__ == '__main__':  
    ##
    # 宣告使用的格式:
    # 定義表頭的樣式
    style_head = {
        "border": Border(left=Side(style='medium', color='FF000000'), 
                         right=Side(style='medium', color='FF000000'),
                         top=Side(style='medium', color='FF000000'), 
                         bottom=Side(style='medium', color='FF000000')),
        "fill": PatternFill("solid", fgColor="9AFF9A"),
        "font": Font(color="000000", bold=True, name="標楷體", size=14),
        "alignment": Alignment(horizontal="center", vertical="center")
        }
    # 定義表格內容樣式
    style_content = {
        "border": Border(left=Side(style='thin', color='FF000000'), 
                         right=Side(style='thin', color='FF000000'),
                         top=Side(style='thin', color='FF000000'), 
                         bottom=Side(style='thin', color='FF000000')),
        "alignment": Alignment(horizontal='left', vertical='center'),
        "font": Font(name="標楷體", size=14)}
    # 定義表格儲存格的大小，2個欄位~7個欄位
    columns_2 = {1:15, 2:60}
    columns_3 = {1:15, 2:30, 3:30}
    columns_4 = {1:15, 2:20, 3:20, 4:20}
    columns_5 = {1:15, 2:15, 3:15, 4:15, 5:15}
    columns_6 = {1:15, 2:12, 3:12, 4:12, 5:12, 6:12}
    columns_7 = {1:15, 2:10, 3:10, 4:10, 5:10, 6:10, 7:10}
    col_idx = {1:'A', 2:'B', 3:'C', 4:'D', 5:'E', 6:'F', 7:'G'} # 欄位名稱(對應工作表 A~G)
    col_obj = {2:columns_2, 3:columns_3, 4:columns_4, 5:columns_5, 6:columns_6, 7:columns_7} #對應取出字典物件(dictionary object)名稱
    
    # 定義備註樣式
    style_note = {
        "border": Border(left=Side(style='thin', color='FF000000'), 
                         right=Side(style='thin', color='FF000000'),
                         top=Side(style='thin', color='FF000000'), 
                         bottom=Side(style='thin', color='FF000000')),
        "alignment": Alignment(horizontal='left', vertical='center'),
        "font": Font(name="標楷體", size=10)}
    
    ##
    # execute main routine with GUI
    GUI()
    
#
# end of main()
##
