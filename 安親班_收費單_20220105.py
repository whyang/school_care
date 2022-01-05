# -*- coding: utf-8 -*-
"""
Created on Mon Jan  3 14:56:13 2022
@author: whyang

合併版本2 說明：
利用Dataframe整理收費名冊，並產出可列印的收費單
1. 讀入[收費資料.xlsx]: 名冊、收費項目、收費期間 等內容
2. 整理成 [收費單_名冊.xlsx]：每個姓名使用1個工作表
3. 由 收費單_名冊.xlsx 產出 [收費單_列印.xlsx]: 加入表頭(該姓名對應的資料_年級、班級)，每個姓名使用1個工作表
4. 使用pandas、openpyxl、win32com.client套件，產生收費單(名冊:Excel工作表/每位學生)、收費單(列印:PDF 3個學生/頁)
"""
import os
import pandas as pd
import numpy as np
import win32com.client
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
def student_list():
    ###
    # Function: 根據「收費資料」內容，組出「收費單_名冊」
    ##
    
    ##
    # 1. 讀入收費資料(學生名單、收費項目、收費月份、註記事項)
    global df_student, df_item, df_month, df_note # set as global variables
    with pd.ExcelFile('安親班_收費資料.xlsx') as xlsxFile:
        df_student = pd.read_excel(xlsxFile, sheet_name='學生名單', header=0)
        df_item = pd.read_excel(xlsxFile, sheet_name='收費項目', header=0) 
        df_month = pd.read_excel(xlsxFile, sheet_name='收費月份', header=0)
        df_note = pd.read_excel(xlsxFile, sheet_name='註記事項', header=0)
    ##
    # 2. 轉收費單表格(每位學生一張)
    df = df_month.T # 轉置'月份'(由一欄轉為一列)
    df.columns = df.iloc[0] # 將月份設定成資料框表頭(欄位名稱)
    df.drop(df.index[0], inplace=True) # 移除原先第一列'月份'各欄位 
    df = pd.concat([df_item, df], axis=1) # 加入'收費項目'欄，合併月份資料框
    #df.fillna(int(0), inplace=True) # 資料框中的cell的空值，用'整數0'取代

    ##
    # 3. 加入學生名單於收費單表格(每位學生一張)；每位學生一個Excel工作表(所有的收費項目)
    name_list = df_student['姓名']
    df_note_temp = pd.DataFrame()
    df_note_temp['項目/月份'] = df_note['註記'] # 將註記內容放在['項目/月份'](註記位置對齊收費單第一欄)
    with pd.ExcelWriter('收費單_名冊.xlsx') as writer:
        for idx, student in enumerate(name_list, start=0):
            # 將註記內容放在['項目/月份'](註記位置對齊收費單第一欄)
            # 收費單(資料框)第一欄名稱為'項目/月份'
            df_temp = pd.concat([df, df_note_temp], axis=0) # 合併收費表單、註記內容
            
            #將每一個學生的收費單轉成numpy array (context)
            content = df_temp.to_numpy() #轉成numpy array        
            # 放回第一列，放入原來資料框欄位名稱
            # 因為資料框轉成numpy arrany後，少了原先欄位名稱(column lable)這一列
            content = np.insert(content, 0, [df_temp.columns], axis=0) # 加入'欄位名稱'到第一列  
            
            # 將每個學生的收費表，加入表頭(學生年級、班級、姓名)
            content = np.insert(content, 0, [np.NAN], axis=0) # 插入新的第一列(空白列)
            # 取回每位學生對應屬於那一個年級與班級
            tableTitle = df_student.iloc[idx]['年級'] + '年' +  df_student.iloc[idx]['班級'] + '班  ' + student
            content[0, 0] = tableTitle
            
            # 產出 收費單_名冊 excel worksheet，將 NAN 用 ''取代
            df_temp = pd.DataFrame(content) #轉成資料框(dataframe) (學生收費單名冊)        
            df_temp.to_excel(writer, sheet_name=student, na_rep='', header=False, index=False) # cell空值 NAN 用 ''取代
#
# end of student_list()
##

def payment_list():
    ###
    # Function: 根據「收費單_名冊」，製作每位學生之收費單表格(Excel檔案，每位學生一張工作表)
    ##
    name_list = df_student['姓名'] # 取出所有收費學生姓名清單
    with pd.ExcelFile('收費單_名冊.xlsx') as xlsxFile:
        wb = load_workbook(xlsxFile)
        for student in name_list:
            ##
            # 讀入每一個學生的收費單(one excel's worksheet)       
            ws = wb[student]

            ##
            # 合併第一列(年級, 班級, 姓名)儲存格 (根據工作表的最大寬度_欄位數目)
            for idx in range(1, ws.max_column+1, 1):
                ws.cell(row=1, column=idx).border = style_head['border']
                ws.cell(row=1, column=idx).fill = style_head['fill']
                ws.cell(row=1, column=idx).font = style_head['font']
                ws.cell(row=1, column=idx).alignment = style_head['alignment']
            ws.row_dimensions[1].height = 30
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)               
        
            ##    
            # 畫出表格(收費表格)，設定格式(style)
            row, col = df_month.shape # 取出 收費表格 的大小(長、寬)(col, row)
            row += 1 # 因為df_month 資料框不含表頭(欄位名稱)，但是收費表格中"欄位名稱"需要顯示為一列
            for j in range(2, row+2, 1): # row 需加上第一列，且range計數結束條件值方式，所以row加上2
                for k in range(1, ws.max_column+1, 1): # 設定每個儲存格(cell)的格式
                    ws.cell(row=j, column=k).border = style_head['border']
                    ws.cell(row=j, column=k).font = style_content['font']
                    ws.cell(row=j, column=k).alignment = style_content['alignment']
                ws.column_dimensions['A'].width = 33
                ws.row_dimensions[j].height = 22
            
            ##
            # 註記事項 
            for idx in range(row+2, ws.max_row+1, 1):
                ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=ws.max_column)
                #ws.cell(row=idx, column=ws.max_column).border = style_content['border']
                ws.cell(row=idx, column=1).font = style_content['font']
                ws.cell(row=idx, column=1).alignment = style_content['alignment']
                ws.row_dimensions[idx].height = 15
        ##
        # 整理所有學生收費單(worksheet)後，儲存到收費單檔案(Excel)
        wb.save('收費單.xlsx') 
        wb.close()     
#
# end of payment_list()
##

def single_payment_list(row_base=0, sourceFile='收費單_名冊.xlsx', targetFile='收費單.xlsx'):
    ###
    # Function: 根據「收費單_名冊」，製作每位學生之收費單表格(Excel檔案，每位學生一張工作表)
    ##
    name_list = df_student['姓名'] # 取出所有收費學生姓名清單
    with pd.ExcelFile(sourceFile) as xlsxFile:
        wb = load_workbook(xlsxFile)
        for student in name_list:
            ##
            # 讀入每一個學生的收費單(one excel's worksheet)       
            ws = wb[student]

            ##
            # 合併第一列(年級, 班級, 姓名)儲存格 (根據工作表的最大寬度_欄位數目)
            i = row_base + 1
            for idx in range(1, ws.max_column+1, 1):
                ws.cell(row=i, column=idx).border = style_head['border']
                ws.cell(row=i, column=idx).fill = style_head['fill']
                ws.cell(row=i, column=idx).font = style_head['font']
                ws.cell(row=i, column=idx).alignment = style_head['alignment']
            ws.row_dimensions[i].height = 30
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=ws.max_column)               
        
            ##    
            # 畫出表格(收費表格)，設定格式(style)
            row, col = df_month.shape # 取出 收費表格 的大小(長、寬)(col, row)
            row += 1 # 因為df_month 資料框不含表頭(欄位名稱)，但是收費表格中"欄位名稱"需要顯示為一列
            for j in range(2, row+2, 1): # row 需加上第一列，且range計數結束條件值方式，所以row加上2
                r = row_base + j
                for k in range(1, ws.max_column+1, 1): # 設定每個儲存格(cell)的格式
                    ws.cell(row=r, column=k).border = style_head['border']
                    ws.cell(row=r, column=k).font = style_content['font']
                    ws.cell(row=r, column=k).alignment = style_content['alignment']
                ws.column_dimensions['A'].width = 33
                ws.row_dimensions[r].height = 22
            
            ##
            # 註記事項 
            for idx in range(row+2, ws.max_row+1, 1): # row 需加上第1列，且range計數起始值設成1開始，所以row加上2
                i = row_base + idx
                ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=ws.max_column)
                ws.cell(row=i, column=1).font = style_content['font']
                ws.cell(row=i, column=1).alignment = style_content['alignment']
                ws.row_dimensions[i].height = 15
        ##
        # 整理所有學生收費單(worksheet)後，儲存到收費單檔案(Excel)
        wb.save(targetFile)
        wb.close()     
#
# end of single_payment_list()
##

def merged_payment_list(noMerge=3, sourceFile='收費單.xlsx', targetFile='收費單_合併.xlsx'):
    ###
    # 程式 2-1
    ##        

    wb_source = load_workbook(sourceFile)
    name_list = df_student['姓名'] # 取出所有收費學生姓名清單   
    length = len(name_list)
    itemRow, itemCol = df_item.shape # number of the table of the item (row, column)
    
    global merged_sheetName
    merged_sheetName = [] # record the whole sheet names after merging

    # 根據 noMerge 值，分群(noMerge+1個新的工作表)
    base = 0
    for i in range(length//noMerge):
        base = i * noMerge # base index of worksheet
        rowIdx = 0 # inital row index of the new worksheet 
        # 合併 noMerge 個工作表
        for idx in range(noMerge):
            sheetName = name_list[base+idx]
            ws = wb_source[sheetName] # 每位學生收費單(工作表)
            if idx == 0:
                sheetName_new = '收費單_列印_'+str(i) 
                merged_sheetName.append(sheetName_new)
                wb_source.create_sheet(sheetName_new)
                ws_target = wb_source[sheetName_new]

            # copy each row in worksheet to the merged target worksheet     
            for index_r, row in enumerate(ws.rows, start=1):
                for index_c, col in enumerate(row, start=1):
                    cell = col.value
                    ws_target.cell(row=rowIdx+index_r, column=index_c).value = cell
                ###
                # editing table
                if index_r == 1:
                    # merge cells in according to each sheet of the tablet
                    # 年級, 班級, 姓名
                    ws_target.merge_cells(start_row=rowIdx+1, start_column=1, end_row=rowIdx+1, end_column=index_c)
                    cellIdx = 'A' + str(rowIdx+1)
                    #ws_target[cellIdx].border = style_head['border']
                    ws_target[cellIdx].fill = style_head['fill']
                    ws_target[cellIdx].font = style_head['font']
                    ws_target[cellIdx].alignment = style_head['alignment']    
                    ws_target.row_dimensions[rowIdx+1].height = 30
                elif (index_r >= 2) and (index_r <= itemRow+2):
                    # 畫表格
                    for k in range(index_c):
                        ws_target.cell(row=rowIdx + index_r, column=k+1).border = style_head['border']
                        ws_target.cell(row=rowIdx + index_r, column=k+1).font = style_content['font']
                        ws_target.cell(row=rowIdx + index_r, column=k+1).alignment = style_content['alignment']
                    ws_target.column_dimensions['A'].width = 33
                    ws_target.row_dimensions[rowIdx + index_r].height = 22
                else:
                    # 註記事項
                    ws_target.merge_cells(start_row=rowIdx+index_r, start_column=1, end_row=rowIdx+index_r, end_column=index_c)
                    #ws_target.cell(row=rowIdx+index_r, column=1).border = style_content['border']
                    ws_target.cell(row=rowIdx+index_r, column=1).font = style_content['font']
                    ws_target.cell(row=rowIdx+index_r, column=1).alignment = style_content['alignment']
                    ws_target.row_dimensions[rowIdx+index_r].height = 15                            
            # update row index
            rowIdx = rowIdx + index_r+1
            # remove single worksheet of each student
            wb_source.remove(wb_source[sheetName])
        
    base = (i+1)*noMerge # base index of worksheet
    rowIdx = 0 # row index of worksheet
    #print(base, rowIdx)
    # 根據 noMerge 值，分群(noMerge+1個新的工作表)最後1個工作表
    for idx in range(length%noMerge):
        sheetName = name_list[base+idx]
        ws = wb_source[sheetName]
        if idx == 0:
            sheetName_new = '收費單_列印_'+str(i+1) 
            merged_sheetName.append(sheetName_new)
            wb_source.create_sheet(sheetName_new)
            ws_target = wb_source[sheetName_new]

        # copy each row in worksheet to the target merged workshee     
        for index_r, row in enumerate(ws.rows, start=1):
            #print(row, index_r)
            for index_c, col in enumerate(row, start=1):
                #print(col, index_c)
                x1 = col.value
                ws_target.cell(row=rowIdx + index_r, column=index_c).value = x1

            ###
            # editing table
            ##           
            if index_r == 1:
                # merge cells according to the table of each sheet
                # 年級, 班級, 姓名
                ws_target.merge_cells(start_row=rowIdx+1, start_column=1, end_row=rowIdx+1, end_column=index_c)
                cellIdx = 'A' + str(rowIdx+1)
                #ws_target[cellIdx].border = style_head['border']
                ws_target[cellIdx].fill = style_head['fill']
                ws_target[cellIdx].font = style_head['font']
                ws_target[cellIdx].alignment = style_head['alignment']    
                ws_target.row_dimensions[rowIdx+1].height = 30
            elif (index_r >= 2) and (index_r <= itemRow+2):
                # 畫表格
                for k in range(index_c):
                    ws_target.cell(row=rowIdx + index_r, column=k+1).border = style_head['border']
                    ws_target.cell(row=rowIdx + index_r, column=k+1).font = style_content['font']
                    ws_target.cell(row=rowIdx + index_r, column=k+1).alignment = style_content['alignment']
                ws_target.column_dimensions['A'].width = 33
                ws_target.row_dimensions[rowIdx + index_r].height = 22
            else:
                # 註記事項
                ws_target.merge_cells(start_row=rowIdx+index_r, start_column=1, end_row=rowIdx+index_r, end_column=index_c)
                #ws_target.cell(row=rowIdx+index_r, column=1).border = style_content['border']
                ws_target.cell(row=rowIdx+index_r, column=1).font = style_content['font']
                ws_target.cell(row=rowIdx+index_r, column=1).alignment = style_content['alignment']
                ws_target.row_dimensions[rowIdx+index_r].height = 15 
        
        # update row index
        rowIdx = rowIdx + index_r + 1 # adding one row used as separation between two student paymentsheets 
        # remove single worksheet of each student
        wb_source.remove(wb_source[sheetName])
    # finally
    wb_source.save(targetFile)
    #wb_source.save('收費單_合併.xlsx')
    wb_source.close()
#
# end of merged_payment_list()
##

def convert_to_pdf():
    ###
    # convet to pdf format
    ##
    o = win32com.client.Dispatch("Excel.Application")
    o.Visible = True #False

    dirpath = os.getcwd()
    filename = '收費單_合併.xlsx'
    wb_p_path = os.path.join(dirpath, filename)

    wb_p = o.Workbooks.Open(wb_p_path)
    ws_index_list = []
    for i, sheetName in enumerate(merged_sheetName, start=1):
        ws_index_list.append(i)

    filename = '收費單_合併.pdf'
    path_to_pdf = os.path.join(dirpath, filename)
    
    wb_p.Sheets(ws_index_list).Select()
    wb_p.ActiveSheet.ExportAsFixedFormat(0, path_to_pdf)
    wb_p.Close(True) 
    o.Quit()
#
# end of convert_to_pdf()
##

if __name__ == '__main__':  
    # 定義表頭的樣式
    style_head = {
        "border": Border(left=Side(style='medium', color='FF000000'), 
                     right=Side(style='medium', color='FF000000'),
                     top=Side(style='medium', color='FF000000'), 
                     bottom=Side(style='medium', color='FF000000')),
        "fill": PatternFill("solid", fgColor="9AFF9A"),
        "font": Font(color="000000", bold=True, name="標楷體", size=14),
        "alignment": Alignment(horizontal="center", vertical="center")
        }

    # 定義表內容樣式
    style_content = {
        "border": Border(left=Side(style='thin', color='FF000000'), 
                     right=Side(style='thin', color='FF000000'),
                     top=Side(style='thin', color='FF000000'), 
                     bottom=Side(style='thin', color='FF000000')),
        "alignment": Alignment(horizontal='left', vertical='center'),
        "font": Font(name="標楷體")}

    ##
    # construct payment lis
    student_list()
    single_payment_list(sourceFile='收費單_名冊.xlsx', targetFile='收費單.xlsx')
    merged_payment_list(noMerge=3, sourceFile='收費單.xlsx', targetFile='收費單_合併.xlsx')
    convert_to_pdf()

    ###
    # 刪除不必要的檔案
    ##
    file_1 = r'收費單_名冊.xlsx'

    try:
        os.remove(file_1)
    except OSError as e:
        print(e)
    else:
        print("Files are deleted successfully")
  
    print('work down')
#
# end of main()
##